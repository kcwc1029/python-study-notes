# ch3_20.py
import openpyxl

fn = 'data3_16.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active                               
for cell in ws['A']:    # A欄
    print(cell.value)
for cell in ws[5]:      # 索引是5
    print(cell.value, end=' ')    


