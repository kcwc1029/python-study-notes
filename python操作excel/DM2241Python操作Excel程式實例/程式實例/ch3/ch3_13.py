# ch3_13.py
import openpyxl

fn = 'data3_11.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active                                   
for i in range(2,ws.max_row+1):                 
    print(ws.cell(row=i,column=2).value)  




