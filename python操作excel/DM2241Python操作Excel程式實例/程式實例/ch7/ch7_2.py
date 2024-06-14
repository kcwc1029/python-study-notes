# ch7_2.py
import openpyxl
from openpyxl.styles import Font, Alignment

wb = openpyxl.Workbook()
ws = wb.active

ws.merge_cells(start_row=1, end_row=2,
               start_column=1, end_column=2)
ws['A1'].font = Font(name='Old English Text MT',
                     color='0000FF',
                     size=20)
ws['A1'].alignment = Alignment(horizontal='center',
                               vertical='center')
ws['A1'] = 'DeepMind'                      
wb.save("out7_2.xlsx")




