# ch6_7.py
import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.Workbook()
ws = wb.active

ws['B2'].fill = PatternFill(fill_type='lightGray',
                            fgColor="0000FF")
ws['B4'].fill = PatternFill(fill_type='lightGray',
                            bgColor="0000FF")
ws['B6'].fill = PatternFill(fill_type='lightGray',
                            fgColor="FF00FF",
                            bgColor="FFFF00")
ws['B8'].fill = PatternFill(patternType='lightGray',
                            fgColor="FFFF00",
                            bgColor="FF00FF")
# 也可以用start_color和end_color
ws['B10'].fill = PatternFill(patternType='lightGray',
                             start_color="FFFF00",
                             end_color="FF00FF")
wb.save("out6_7.xlsx")




