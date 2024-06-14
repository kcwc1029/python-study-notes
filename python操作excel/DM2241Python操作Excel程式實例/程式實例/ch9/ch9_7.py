# ch9_7.py
import openpyxl

fn = "data9_7.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['C7'] = "=SUM(C4:C6)"
ws['C8'] = "=MAX(C4:C6)"
ws['C9'] = "=MIN(C4:C6)"
wb.save("out9_7.xlsx")











