# ch6_5.py
import openpyxl
from openpyxl.styles import Border, Side

fn = "data6_5.xlsx"
wb = openpyxl.load_workbook(fn,data_only=True)
ws = wb.active
data = ws[2]                        # 薪資項目
row = 0;
for i in range(4,8):
    ws.insert_rows(i+row,2)         # 插入 2 列空白列               
    for j, d in enumerate(data):    # 取得元組內容         
        ws.cell(row=i+row+1,column=j+1,value=d.value)
    row = row + 2

side = Side(border_style='thin')
borders = Border(left=side,right=side,top=side,bottom=side)
for rows in ws['A4':'H14']:
    for cell in rows:
        cell.border = borders 
wb.save("out6_5.xlsx")




