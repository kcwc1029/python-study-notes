# ch4_5.py
import openpyxl 

fn1 = "data4_5.xlsx"                    # 來源活頁簿
wb = openpyxl.load_workbook(fn1)
fn2 = "dst4_5.xlsx"
new_wb = openpyxl.load_workbook(fn2)    # 建立目的活頁簿
for i in range(4):
    ws = wb.worksheets[i]
    dst_title = ws.title    
    new_ws = new_wb.create_sheet(title=dst_title)
    for data in ws.iter_rows(min_row=1,max_row=ws.max_row,
            min_col=1,max_col=ws.max_column, values_only=True):
        value = list(data)
        new_ws.append(value)            # 寫入目的活頁簿    
new_wb.save("out4_5.xlsx")              # 儲存結果





