# ch6_7_2.py
from openpyxl import Workbook
from openpyxl.styles import GradientFill    

wb = Workbook()
ws = wb.active
# 2 個顏色的線性填滿
ws['B2'].fill = GradientFill(type='linear',stop=("FFFF00","00FF00"))
# 3 個顏色的線性填滿
ws['D2'].fill = GradientFill(type='linear',
                             stop=("FF0000","0000FF","00FF00"))
# 3 個顏色的線性填滿, 45度旋轉
ws['F2'].fill = GradientFill(type='linear',
                             stop=("FF0000","0000FF","00FF00"),degree=45)
# 3 個顏色的線性填滿, 90度旋轉
ws['H2'].fill = GradientFill(type='linear',
                             stop=("FF0000","0000FF","00FF00"),degree=90)
# 3 個顏色的線性填滿, 135度旋轉
ws['J2'].fill = GradientFill(type='linear',
                            stop=("FF0000","0000FF","00FF00"),degree=135)
wb.save('out6_7_2.xlsx')


