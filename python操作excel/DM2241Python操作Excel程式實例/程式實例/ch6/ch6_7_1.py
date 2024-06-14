# ch6_7_1.py
import openpyxl
from openpyxl.styles import PatternFill

fn = "data6_7_1.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

for rows in ws.iter_rows(min_row=1,max_row=6,min_col=1,max_col=8):
    for cell in rows:
        if cell.row % 2:
            cell.fill = PatternFill(start_color="FFFF00",
                                    fill_type="solid")
wb.save("out6_7_1.xlsx")




