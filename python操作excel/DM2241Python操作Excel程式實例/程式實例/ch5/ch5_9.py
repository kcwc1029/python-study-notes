# ch5_9.py
import openpyxl

fn = "data5_8.xlsx"
wb = openpyxl.load_workbook(fn,data_only=True)
ws = wb.active
ws.insert_cols(3,2)
ws['C3'] = 'ID'
ws['D3'] = '性別'
wb.save("out5_9.xlsx")




