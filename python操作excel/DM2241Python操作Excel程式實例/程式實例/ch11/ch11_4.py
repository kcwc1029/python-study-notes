# ch11_4.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_4.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義儲存格格式
font = Font(bold=True,color='FFFFFF')           # 字型
bgRed = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgRed)
# 應用姓 洪 業務員的資料
rule = Rule(type='beginsWith', operator='beginsWith',
            formula=['LEFT(A1,1)="洪"'],
            dxf=dxf)
ws.conditional_formatting.add('A1:F10',rule)
# 儲存結果
wb.save('out11_4.xlsx')


