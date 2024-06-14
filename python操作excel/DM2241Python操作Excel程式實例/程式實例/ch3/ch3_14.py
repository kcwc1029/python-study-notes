# ch3_14.py
import openpyxl

fn = 'data3_14.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active
for i in range(1,ws.max_row+1):         # row做索引增值
    for j in range(1,ws.max_column+1):  # column做索引增值
        print(f"{ws.cell(row=i,column=j).value}", end=" ")   
    print()                             # 換列輸出


  




