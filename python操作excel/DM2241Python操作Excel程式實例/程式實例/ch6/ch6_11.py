# ch6_11.py
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import Color

fn = "data6_1.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['B1'].font = Font(color=Color(indexed=6))
ws['B2'].font = Font(underline='single')
ws['C2'].font = Font(underline='double')
ws['B3'].font = Font(color=Color(indexed=40),
                     italic=True)                         
wb.save("out6_11.xlsx")




