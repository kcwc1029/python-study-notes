# ch7_16.py
from openpyxl import Workbook
from openpyxl.styles import GradientFill    

wb = Workbook()
ws = wb.active

# top=0.0
ws.merge_cells('B2:C5')
ws['B2'].fill = GradientFill(type='path',top="0.0",
                             stop=("00FF00","FFFF00"))
# top=0.2
ws.merge_cells('E2:F5')
ws['E2'].fill = GradientFill(type='path',top="0.2",
                             stop=("00FF00","FFFF00"))
# top=0.5
ws.merge_cells('H2:I5')
ws['H2'].fill = GradientFill(type='path',top="0.5",
                             stop=("00FF00","FFFF00"))
# top=0.8
ws.merge_cells('B7:C10')
ws['B7'].fill = GradientFill(type='path',top="0.8",
                             stop=("00FF00","FFFF00"))
# top=1.0
ws.merge_cells('E7:F10')
ws['E7'].fill = GradientFill(type='path',top="1.0",
                             stop=("00FF00","FFFF00"))
wb.save('out7_16.xlsx')


