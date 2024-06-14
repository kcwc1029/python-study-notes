# ch11_2.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_1.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義介於50和60分之間的儲存格格式
font = Font(bold=True) 
bgRed = PatternFill(start_color='FFFF00',
                    end_color='FFFF00',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgRed)
# 應用介於50和60分之間的資料
rule = Rule(type='cellIs',operator='between',
            formula=[50,59],dxf=dxf)
ws.conditional_formatting.add('A2:C11',rule)

# 儲存結果
wb.save('out11_2.xlsx')


