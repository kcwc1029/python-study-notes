# ch8_10.py
import openpyxl
import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.column_dimensions['B'].width = 40
ws['B2'] = datetime.datetime.today()
ws['B3'] = datetime.datetime.today()
ws['B3'].number_format = 'yyyy-mm-dd hh:mm:ss'
ws['B4'] = datetime.datetime.today()
ws['B4'].number_format = 'yyyy年mm月dd日 hh時mm分ss秒'
wb.save("out8_10.xlsx")






