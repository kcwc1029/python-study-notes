# ch3_2.py
import openpyxl

fn = 'data3_2.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active                       
print("儲存格B2 = ", ws['B2'].value)    # B2 
print("儲存格B3 = ", ws['B3'].value)    # B3
print("儲存格B4 = ", ws['B4'].value)    # B4
print("儲存格C3 = ", ws['C3'].value)    # C3
print("儲存格C4 = ", ws['C4'].value)    # C4


