# ch11_9_1.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_1.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義儲存格格式
font = Font(bold=True,color='FFFFFF')           # 字型
bgBlue = PatternFill(start_color='0000FF',
                    end_color='0000FF',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgBlue)
# 應用top10的資料
rule1 = Rule(type='top10',rank=10,percent=True,dxf=dxf)
ws.conditional_formatting.add('A2:A11',rule1)
# 應用top 3的資料
rule2 = Rule(type='top10',rank=3,dxf=dxf)
ws.conditional_formatting.add('B2:B11',rule2)
# 儲存結果
wb.save('out11_9_1.xlsx')


