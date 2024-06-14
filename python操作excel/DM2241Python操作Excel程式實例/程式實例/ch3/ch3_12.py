# ch3_12.py
import openpyxl

fn = 'data3_11.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active                                   
for i in range(2,ws.max_column+1):                 
    print(ws.cell(row=3,column=i).value, end=' ')  




