# ch9_11.py
import openpyxl
from openpyxl.formula.translate import Translator

fn = "data9_6.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['E4'] = "=RANK(D4,$D$4:$D$9)"
for i in range(5,10):
    index = 'E' + str(i)
    ws[index] = Translator("=RANK(D4,$D$4:$D$9)",
                          origin="E4").translate_formula(index)
wb.save("out9_11.xlsx")











