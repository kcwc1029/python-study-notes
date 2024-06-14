# ch3_11.py
import openpyxl

fn = 'data3_11.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active
print(f"工作表欄數 = {ws.max_column}")    
print(f"工作表列數 = {ws.max_row}")    



