# ch6_8.py
import openpyxl
from openpyxl.styles import Alignment

fn = "data6_8.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['B2'].alignment = Alignment(horizontal='right')  
ws['C2'].alignment = Alignment(horizontal='center') 
ws['D2'].alignment = Alignment(horizontal='left')
ws['F2'].alignment = Alignment(horizontal='centerContinuous')
ws['B4'].alignment = Alignment(text_rotation=30)  
ws['C4'].alignment = Alignment(text_rotation=45) 
ws['D4'].alignment = Alignment(text_rotation=60)
wb.save("out6_8.xlsx")




