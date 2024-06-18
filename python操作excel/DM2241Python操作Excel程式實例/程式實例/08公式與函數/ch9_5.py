# ch9_5.py
import openpyxl

fn = "data9_5.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['C3'] = "=TODAY()"
ws['C3'].number_format = 'yyyy/m/d'
ws['E6'] = '=DATEDIF(D6,$C$3,"Y")'
ws['F6'] = '=DATEDIF(D6,$C$3,"YM")'
ws['G6'] = '=DATEDIF(D6,$C$3,"MD")'
wb.save("out9_5.xlsx")











