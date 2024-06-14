# ch11_1.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_1.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義低於60分的儲存格格式
font = Font(bold=True,color='FFFFFF')           # 字型
bgRed = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgRed)
# 應用低於60分的資料
rule = Rule(type='cellIs',operator='lessThan',
            formula=[60],dxf=dxf)
ws.conditional_formatting.add('A2:C11',rule)

# 定義大於或等於80分的儲存格格式
font = Font(bold=True,color='0000FF')           # 字型
dxf = DifferentialStyle(font=font)
# 應用大於或等於80分的資料
rule = Rule(type='cellIs',operator='greaterThanOrEqual',
            formula=[80],dxf=dxf)
ws.conditional_formatting.add('A2:C11',rule)
# 儲存結果
wb.save('out11_1.xlsx')


