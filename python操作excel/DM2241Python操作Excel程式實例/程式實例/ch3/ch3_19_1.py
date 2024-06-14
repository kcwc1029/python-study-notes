# ch3_19_1.py
import openpyxl

fn = 'data3_19_1.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active
print(f"工作表有資料最小列數 = {ws.min_row}")
print(f"工作表有資料最大列數 = {ws.max_row}")
print(f"工作表有資料最小欄數 = {ws.min_column}")
print(f"工作表有資料最大欄數 = {ws.max_column}")






