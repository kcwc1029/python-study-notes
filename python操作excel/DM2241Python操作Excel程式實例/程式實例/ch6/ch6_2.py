# ch6_2.py
import openpyxl
from openpyxl.styles import Font

fn = "data6_1.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['B1'].font = Font(color='0000FF')
ws['B2'].font = Font(underline='single')
ws['C2'].font = Font(underline='double')
for i in range(3,7):
    ws['B'+str(i)].font = Font(color='0000FF',
                               italic=True)                         
wb.save("out6_2.xlsx")




