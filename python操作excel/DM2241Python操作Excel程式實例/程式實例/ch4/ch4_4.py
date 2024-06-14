# ch4_4.py
import openpyxl 

fn = "data4_2.xlsx"                     # 來源活頁簿
wb = openpyxl.load_workbook(fn)
ws = wb.active
dst = "data4_4.xlsx"
new_wb = openpyxl.load_workbook(dst)    # 開啟目的活頁簿
new_ws = new_wb.create_sheet(title="新SPA客戶表")
for data in ws.iter_rows(min_row=1,max_row=ws.max_row,
            min_col=1,max_col=ws.max_column, values_only=True):
    value = list(data)
    new_ws.append(value)        # 寫入目的活頁簿

new_wb.save("out4_4.xlsx")      # 用新活頁簿儲存結果





