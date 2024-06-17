# ch1_2_1.py
import openpyxl

fn = 'sales.xlsx'
wb = openpyxl.load_workbook(fn)     # wb是Excel檔案物件    
print("所有工作表     = ", wb.sheetnames)
for sheet in wb.sheetnames:
    print("工作表名稱 = ", sheet)






