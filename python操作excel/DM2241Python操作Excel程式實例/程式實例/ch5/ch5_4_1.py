# ch5_4_1.py
import openpyxl

fn = "data5_1.xlsx"
wb = openpyxl.load_workbook(fn,data_only=True)
ws = wb.active
area = ws['A2':'H2']                    # 薪資項目
row = 0;
for i in range(4,8):
    ws.insert_rows(i+row,2)             # 插入 2 列空白列
    for datarow in area:
        data = list(datarow)            # 轉成串列
        for j, d in enumerate(data):    # 取得串列內容         
            ws.cell(row=i+row+1,column=j+1,value=d.value)
    row = row + 2
wb.save("out5_4_1.xlsx")




