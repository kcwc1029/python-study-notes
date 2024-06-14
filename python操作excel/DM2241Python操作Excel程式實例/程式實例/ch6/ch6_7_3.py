# ch6_7_3.py
from openpyxl import Workbook
from openpyxl.styles import GradientFill    

wb = Workbook()
ws = wb.active

ws.merge_cells('B2:C4')
ws['B2'].fill = GradientFill(type='path',top="0.0",stop=("00FF00","FF0000"))

ws.merge_cells('E2:F4')
# top=0.2
ws['E2'].fill = GradientFill(type='path',top="0.2",stop=("00FF00","FF0000"))

ws.merge_cells('H2:I4')
# top=0.5
ws['H2'].fill = GradientFill(type='path',top="0.5",stop=("00FF00","FF0000"))

ws.merge_cells('B6:C8')
# top=0.8
ws['B6'].fill = GradientFill(type='path',top="0.8",stop=("00FF00","FF0000"))

ws.merge_cells('E6:F8')
# top=1.0
ws['E6'].fill = GradientFill(type='path',top="1.0",stop=("00FF00","FF0000"))



wb.save('out6_7_3.xlsx')


