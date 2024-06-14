# ch3_9.py
import openpyxl

fn = 'data3_6.xlsx'
wb = openpyxl.load_workbook(fn, data_only=True)
ws = wb.active
print("儲存格B4 = ", ws.cell(row=4, column=2).value)     
print("儲存格B5 = ", ws.cell(row=5, column=2).value)    
print("儲存格B6 = ", ws.cell(row=6, column=2).value)    
print("儲存格B7 = ", ws.cell(row=7, column=2).value)     
print("儲存格B8 = ", ws.cell(row=8, column=2).value)
print("儲存格C4 = ", ws.cell(row=4, column=2).value)
print("儲存格C5 = ", ws.cell(row=5, column=3).value)    
print("儲存格C6 = ", ws.cell(row=6, column=3).value)
print("儲存格C7 = ", ws.cell(row=7, column=3).value)   
print("儲存格C8 = ", ws.cell(row=8, column=3).value)    


