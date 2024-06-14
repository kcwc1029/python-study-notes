# ch3_1.py
import openpyxl

wb = openpyxl.Workbook()    # 建立空白的活頁簿
ws = wb.active              # 取得目前工作表
ws['A2'] = 'Apple'
ws['A3'] = 'Orange'
ws['B2'] = 200
ws['B3'] = 150
wb.save('out3_1.xlsx')      # 將活頁簿儲存





