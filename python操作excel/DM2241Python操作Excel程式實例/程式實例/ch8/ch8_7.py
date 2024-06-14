# ch8_7.py
import openpyxl

fn = "data8_7.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['B2'].number_format = 'm/d/yy'
ws['B3'].number_format = 'mm-dd-yyyy'
ws['B4'].number_format = 'yyyy-mm-dd'
ws['B5'].number_format = 'd-mmm-yy'
ws['B6'].number_format = 'h:mm AM/PM'
ws['B7'].number_format = 'h:mm'
wb.save("out8_7.xlsx")



