# ch4_6.py
import openpyxl 

fn = "data4_5.xlsx"                 # 來源活頁簿
wb = openpyxl.load_workbook(fn)
ws = wb.active
for i in range(4):
    ws = wb.worksheets[i]
    fname = ws.title
    new_wb = openpyxl.Workbook()    # 建立目的的活頁簿
    new_ws = new_wb.active
    for data in ws.iter_rows(min_row=1,max_row=ws.max_row,
            min_col=1,max_col=ws.max_column, values_only=True):
        value = list(data)
        new_ws.append(value)        # 寫入目的活頁簿
    fname = fname + '.xlsx'
    new_wb.save(fname)              # 儲存結果





