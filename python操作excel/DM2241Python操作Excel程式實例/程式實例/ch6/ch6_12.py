# ch6_12.py
import openpyxl
from openpyxl.styles import Font, NamedStyle, Border, Side

fn = "data6_12.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

namestyle = NamedStyle(name="nameSample")
namestyle.font = Font(bold=True,color="0000FF")
bd = Side(style='thick', color="FF0000")
namestyle.border = Border(left=bd, top=bd, right=bd, bottom=bd)
wb.add_named_style(namestyle)   # 註冊
ws['B2'].style = namestyle      # 用法 1, 直接使用
ws['B4'].style = 'nameSample'   # 用法 2
wb.save("out6_12.xlsx")




