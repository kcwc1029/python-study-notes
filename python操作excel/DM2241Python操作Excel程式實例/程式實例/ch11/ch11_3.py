# ch11_3.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_3.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義低於平均成績的儲存格格式
font = Font(bold=True,color='FFFFFF')           # 字型
bgRed = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgRed)
# 應用低於平均成績的資料
rule = Rule(type='cellIs',operator='lessThan',
            formula=['$E$3'],dxf=dxf)
ws.conditional_formatting.add('A2:C11',rule)

# 定義大於或等於平均成績的儲存格格式
font = Font(bold=True,color='FFFFFF')           # 字型
bgBlue = PatternFill(start_color='0000FF',
                     end_color='0000FF',
                     fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgBlue)
# 應用大於或等於平均成績的資料
rule = Rule(type='cellIs',operator='greaterThanOrEqual',
            formula=['$E$3'],dxf=dxf)
ws.conditional_formatting.add('A2:C11',rule)

# 儲存結果
wb.save('out11_3.xlsx')


