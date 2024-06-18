# ch9_8.py
import openpyxl
from openpyxl.utils import get_column_letter

fn = "data9_3.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
for i in range(3,7):
    ch = get_column_letter(i)       # 將數字轉成欄位
    index = ch + str(7)
    start_index = ch + str(4)
    end_index = ch + str(6)
    ws[index] = "=SUM({}:{})".format(start_index,end_index)
wb.save("out9_8.xlsx")












