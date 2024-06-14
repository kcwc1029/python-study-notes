# ch3_10.py
import openpyxl

fn = 'data3_10.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active
print(f"A1 = {ws['A1'].value}")
print(f"A1 = {ws['A1'].column}, {ws['A1'].row}, {ws['A1'].coordinate}")
print(f"A2 = {ws['A2'].value}")
print(f"A2 = {ws['A2'].column}, {ws['A2'].row}, {ws['A2'].coordinate}")
print(f"A3 = {ws['A3'].value}")
print(f"A3 = {ws['A3'].column}, {ws['A3'].row}, {ws['A3'].coordinate}")
print(f"B1 = {ws['B1'].value}")
print(f"B1 = {ws['B1'].column}, {ws['B1'].row}, {ws['B1'].coordinate}")
print(f"B2 = {ws['B2'].value}")
print(f"B2 = {ws['B2'].column}, {ws['B2'].row}, {ws['B2'].coordinate}")    
print(f"B3 = {ws['B3'].value}")
print(f"B3 = {ws['B3'].column}, {ws['B3'].row}, {ws['B3'].coordinate}")  



