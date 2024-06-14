# ch3_4.py
import openpyxl

fn = 'data3_2.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active                       
print("儲存格B2 = ", ws.cell(row=2, column=2).value)    # B2
print("儲存格B3 = ", ws.cell(row=3, column=2).value)    # B3
print("儲存格B4 = ", ws.cell(row=4, column=2).value)    # B4
print("儲存格C3 = ", ws.cell(row=3, column=3).value)    # C3
print("儲存格C4 = ", ws.cell(row=4, column=3).value)    # C4



