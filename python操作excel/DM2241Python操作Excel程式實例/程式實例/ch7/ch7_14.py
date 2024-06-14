# ch7_14.py
import openpyxl
from openpyxl.styles import Protection

fn = "data7_14.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws.protection.enable()              
ws.merge_cells('B4:C8')
ws['B4'].protection = Protection(locked=False, hidden=False)
for i in range(4,9):
    index = 'E' + str(i) + ':' + 'G' + str(i)
    ws.merge_cells(index)
    index = 'E' + str(i) 
    ws[index].protection = Protection(locked=False, hidden=False)
ws.merge_cells('D9:G9')
ws['D9'].protection = Protection(locked=False, hidden=False)
wb.save("out7_14.xlsx")




