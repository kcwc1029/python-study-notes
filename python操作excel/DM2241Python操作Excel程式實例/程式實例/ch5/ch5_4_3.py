# ch5_4_3.py
import openpyxl

fn = "data5_1.xlsx"
wb = openpyxl.load_workbook(fn,data_only=True)
ws = wb.active
print("執行前")
for r in ws.iter_rows(values_only=True):    # 執行前輸出
    print(r)
row = 0;
for i in range(4,8):
    ws.insert_rows(i+row,1)
    row = row + 1
print("執行後")
for r in ws.iter_rows(values_only=True):    # 執行前輸出
    print(r)
wb.save("out5_4_3.xlsx")




