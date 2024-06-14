# ch4_3.py
import openpyxl 

fn = "data4_2.xlsx"             # 來源活頁簿
wb = openpyxl.load_workbook(fn)
ws = wb.active

new_wb = openpyxl.Workbook()    # 建立目的的活頁簿
new_ws = new_wb.active

for data in ws.iter_rows(min_row=1,max_row=ws.max_row,
            min_col=1,max_col=ws.max_column, values_only=True):
    value = list(data)
    new_ws.append(value)        # 寫入目的活頁簿

new_wb.save("out4_3.xlsx")      # 儲存結果





