# ch6_3.py
import openpyxl
from openpyxl.styles import Font, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
# 建立含13種框線樣式的串列
border_styles = ['hair','dotted','dashed','dashDotDot',
                 'dashDot','thin','mediumDashDotDot',
                 'mediumDashDot','slantDashDot','mediumDashed', 
                 'medium','thick','double']

# 建立輸出含13個框線的列號串列
rows = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26]

for row, border_style in zip(rows, border_styles): 
    for col in [2, 4, 6]:   # B, D, F 欄        
        if col == 2:        # 如果是 B 欄, 用藍色輸出框線樣式名稱          
            ws.cell(row=row, column=col).value=border_style
            ws.cell(row=row, column=col).font = Font(color='0000FF')
        elif col == 4:      # 如果是 D 欄, 設定左上至右下的紅色對角線
            side = Side(border_style=border_style, color='FF0000')
                            # 建立左上到右下對角線物件
            diagDown = Border(diagonal=side,diagonalDown=True)
                            # 建立紅色對角線
            ws.cell(row=row, column=col).border = diagDown     
        else:               # 如果是 F 欄, 建立框線
            side = Side(border_style=border_style)
                            # 建立儲存格四周的框線
            borders = Border(left=side,right=side,top=side,bottom=side)
            ws.cell(row=row, column=col).border = borders                               
wb.save("out6_3.xlsx")




