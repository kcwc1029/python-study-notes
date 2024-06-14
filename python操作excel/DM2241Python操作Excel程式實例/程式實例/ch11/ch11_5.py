# ch11_5.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_5.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義低於平均成績的儲存格格式
font = Font(bold=True,color='FFFFFF')           # 字型
bgRed = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgRed)
# 應用低於平均成績的資料
rule = Rule(type='containsText', operator='containsText',
            formula=['NOT(ISERROR(SEARCH("大學",A1)))'],
            dxf=dxf)
ws.conditional_formatting.add('A1:E151',rule)
# 儲存結果
wb.save('out11_5.xlsx')


