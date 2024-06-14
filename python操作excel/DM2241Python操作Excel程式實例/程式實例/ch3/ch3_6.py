# ch3_6.py
import openpyxl

fn = 'data3_6.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active
print("儲存格B4 = ", ws['B4'].value)     
print("儲存格B5 = ", ws['B5'].value)    
print("儲存格B6 = ", ws['B6'].value)    
print("儲存格B7 = ", ws['B7'].value)     
print("儲存格B8 = ", ws['B8'].value)
print("儲存格C4 = ", ws['C4'].value)
print("儲存格C5 = ", ws['C5'].value)    
print("儲存格C6 = ", ws['C6'].value)
print("儲存格C7 = ", ws['C7'].value)    
print("儲存格C8 = ", ws['C8'].value)    


