# ch11_10.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_1.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義儲存格格式 - 背景是藍色
font = Font(bold=True,color='FFFFFF')           # 字型
bgBlue = PatternFill(start_color='0000FF',
                    end_color='0000FF',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgBlue)
# 應用top10的資料
rule1 = Rule(type='top10',rank=10,percent=True,dxf=dxf)
ws.conditional_formatting.add('A2:A11',rule1)
# 應用top30的資料
rule2 = Rule(type='top10',rank=30,percent=True,dxf=dxf)
ws.conditional_formatting.add('B2:B11',rule2)
# 定義儲存格格式 - 背景是紅色
font = Font(bold=True,color='FFFFFF')           # 字型
bgRed = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgRed)
# 應用bottom30的資料
rule3 = Rule(type='top10',rank=30,
             bottom=True,percent=True,dxf=dxf)
ws.conditional_formatting.add('C2:C11',rule3)
# 儲存結果
wb.save('out11_10.xlsx')


