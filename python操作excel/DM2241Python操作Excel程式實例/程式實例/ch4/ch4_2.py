# ch4_2.py
import openpyxl 

fn = "data4_2.xlsx"             # 來源活頁簿
wb = openpyxl.load_workbook(fn)
ws = wb.active

new_wb = openpyxl.Workbook()    # 建立目的的活頁簿
new_ws = new_wb.active

for m in range(1, ws.max_row+1):
    for n in range(65, 65+ws.max_column):   # 65是A
        ch = chr(n)              # 將ASCII碼值轉字元
        index = ch + str(m)
        data =  ws[index].value
        new_ws[index].value = data  # 寫入目的活頁簿

new_wb.save("out4_2.xlsx")          # 儲存結果





