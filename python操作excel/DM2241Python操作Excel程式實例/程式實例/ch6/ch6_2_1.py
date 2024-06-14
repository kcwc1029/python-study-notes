# ch6_2_1.py
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
ws = wb.active
ws['B2'] = "Ming-Chi Institute of Technology"
ws['B2'].font = Font(name='Old English Text MT',color='0000FF')
ws['B4'] = "明志工專"
ws['B4'].font = Font(name='標楷體',color='0000FF')                           
wb.save("out6_2_1.xlsx")




