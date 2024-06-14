# ch11_12.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_1.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義儲存格格式
font = Font(bold=True,color='FFFFFF')           # 字型
bgBlue = PatternFill(start_color='0000FF',
                    end_color='0000FF',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgBlue)
# 應用 aboveAverage 的資料
rule1 = Rule(type='aboveAverage',dxf=dxf)
ws.conditional_formatting.add('A2:A11',rule1)
# 定義儲存格格式
bgRed = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgRed)
# 應用 低於平均 的資料
rule2 = Rule(type='aboveAverage',aboveAverage=False,dxf=dxf)
ws.conditional_formatting.add('B2:B11',rule2)
# 儲存結果
wb.save('out11_12.xlsx')


