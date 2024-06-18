# ch9_10.py
import openpyxl
from openpyxl.formula.translate import Translator

fn = "data9_3.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['C7'] = "=SUM(C4:C6)"
ws['D7'] = Translator("=SUM(C4:C6)",
                       origin="C7").translate_formula("D7")
ws['E7'] = Translator("=SUM(C4:C6)",
                       origin="C7").translate_formula("E7")
ws['F7'] = Translator("=SUM(C4:C6)",
                       origin="C7").translate_formula("F7")
wb.save("out9_10.xlsx")












