# ch7_13.py
import openpyxl
from openpyxl.styles import Protection

wb = openpyxl.Workbook()
ws = wb.active
ws.protection.enable()              
for row in ws['A1:B2']:
    for cell in row:
        cell.protection = Protection(locked=False, hidden=False)
ws.merge_cells('C1:E2')
ws['C1'].protection = Protection(locked=False, hidden=False)
wb.save("out7_13.xlsx")




