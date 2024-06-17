# ch7_17.py
from openpyxl import Workbook
from openpyxl.styles import GradientFill    

wb = Workbook()
ws = wb.active

# top = 0.5, left = 0.5
ws.merge_cells('B2:C5')
ws['B2'].fill = GradientFill(type='path',top="0.5",left="0.5",
                             stop=("0000FF","FFFFFF"))
# top = 0.5, left = 0.5
ws.merge_cells('E2:F5')
ws['E2'].fill = GradientFill(type='path',top="0.5",left="0.5",
                             stop=("FFFFFF","0000FF"))
wb.save('out7_17.xlsx')


