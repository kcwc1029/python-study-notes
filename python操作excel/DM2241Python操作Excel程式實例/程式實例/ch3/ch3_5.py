# ch3_5.py
import openpyxl
import time

wb = openpyxl.Workbook()    # 建立空白的活頁簿
ws = wb.active              # 取得目前工作表
ws['A1'] = time.strftime("%Y/%m/%d")
ws['A2'] = '期貨行情'
ws['A3'] = '小麥'
ws['A4'] = '玉米'
ws['B3'] = 1097
ws['B4'] = 742
wb.save('out3_5.xlsx')      # 將活頁簿儲存





