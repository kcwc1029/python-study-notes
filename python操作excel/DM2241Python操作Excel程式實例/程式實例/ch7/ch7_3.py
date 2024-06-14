# ch7_3.py
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border

fn = "data7_3.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

ws.merge_cells('A1:E1')
ws['A1'].font = Font(color='0000FF')

side = Side(border_style='thin')
borders = Border(left=side,right=side,top=side,bottom=side)
for rows in ws['A1':'E9']:
    for cell in rows:
        cell.border = borders
        cell.alignment = Alignment(horizontal='center')                     
wb.save("out7_3.xlsx")




