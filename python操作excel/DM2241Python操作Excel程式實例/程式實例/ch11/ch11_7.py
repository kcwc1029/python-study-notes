# ch11_7.py
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

fn = "data11_7.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active

# 定義儲存格格式
font = Font(bold=True,color='FFFFFF')           # 字型
bgRed = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
dxf = DifferentialStyle(font=font,fill=bgRed)
# 下個月到期的支票
rule = Rule(type='timePeriod',timePeriod='nextMonth',dxf=dxf)
ws.conditional_formatting.add('B2:B7',rule)
# 儲存結果
wb.save('out11_7.xlsx')


