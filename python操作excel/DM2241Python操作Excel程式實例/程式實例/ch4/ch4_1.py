# ch4_1.py
import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "DataRange"
for row in range(1, 20):
    ws1.append(range(500))
ws2 = wb.create_sheet(title="School")
ws2['F5'] = "明志科技大學"
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        ws3.cell(column=col,row=row,value="{0}".format(get_column_letter(col)))
wb.save("out4_1.xlsx")




