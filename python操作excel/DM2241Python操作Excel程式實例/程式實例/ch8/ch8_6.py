# ch8_6.py
import openpyxl

fn = "data8_6.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.active
ws['B2'].number_format = '#.##'
ws['B3'].number_format = '#.##'
ws['B4'].number_format = '#0.##'
ws['B5'].number_format = '#0.##'
ws['B6'].number_format = '000.00'
ws['B7'].number_format = '#.00'
ws['B8'].number_format = '[Red]#.00'
wb.save("out8_6.xlsx")



