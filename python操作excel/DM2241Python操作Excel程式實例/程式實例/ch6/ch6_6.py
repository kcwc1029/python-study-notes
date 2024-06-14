# ch6_6.py
import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.Workbook()
ws = wb.active

# 建立圖案樣式串列
patterns = ['solid','darkDown','darkGray',
            'darkGrid','darkHorizontal','darkTrellis',
            'darkUp','darkVertical','gray0625',
            'gray125','lightDown','lightGray',
            'lightGrid','lightHorizontal','lightTrellis',
            'lightUp','lightVertical','mediumGray']       

# 設定儲存格區間
cells = ws.iter_cols(min_row=2,max_row=20,min_col=2,max_col=3)
for col in cells:
    for cell, pattern in zip(col,patterns):         
        if cell.col_idx == 2 :  # 如果是 B 欄則輸出圖案樣式
            cell.fill = PatternFill(fill_type=pattern)     
        else:                   # 否則輸出圖案名稱
            cell.value = pattern  
wb.save("out6_6.xlsx")




