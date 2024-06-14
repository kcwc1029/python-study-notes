# ch3_3_1.py
import openpyxl

wb = openpyxl.Workbook()    # 建立空白的活頁簿
ws = wb.active              # 取得目前工作表
ws.cell(row=2, column=1).value = 'Apple'
ws.cell(row=3, column=1).value = 'Orange'
ws.cell(row=2, column=2).value = 200
ws.cell(row=3, column=2).value = 150
wb.save('out3_3_1.xlsx')    # 將活頁簿儲存





