# ch7_10.py
import openpyxl
from openpyxl.comments import Comment

wb = openpyxl.Workbook()
ws = wb.active
ws['B2'] = "楊貴妃"
comment = Comment("唐朝美女","洪錦魁")
ws['B2'].comment = comment
print(f"註解 : {comment.text}")
print(f"作者 : {comment.author}")
comment.width = 250
comment.height = 50
wb.save("out7_10.xlsx")




