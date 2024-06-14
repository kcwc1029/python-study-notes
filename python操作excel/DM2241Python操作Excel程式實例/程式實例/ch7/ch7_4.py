# ch7_4.py
import openpyxl
from openpyxl.styles import Font, Alignment

wb = openpyxl.Workbook()
ws = wb.active

ws['B2'] = "深智"
wb.save("out7_4.xlsx")

ws.merge_cells('B2:C3')
ws['B2'].font = Font(name='Old English Text MT',
                     color='0000FF',)
ws['B2'].alignment = Alignment(horizontal='center',
                               vertical='center')                     
wb.save("out7_4_1.xlsx")

ws.unmerge_cells('B2:C3')
wb.save("out7_4_2.xlsx")




